"""
Загрузка каталога из Excel с поддержкой артикулов для синхронизации с 1C.

Структура Excel:
- Каждый лист = одна категория (Импланты, Протетика, Лаборатория, Наборы, материалы)
- На каждом листе: Категория (опц.) | Линейка | Название товара | Артикул (1C) | Тип | Диаметр | Длина | Высота абатмента | Ед. изм.

Настройка: catalog_config.py — маппинг колонок и структура навигации.

Использование:
    python load_catalog_from_excel.py [filename.xlsx]
"""
import sys
import json
import io
import logging
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

# Устанавливаем UTF-8 для вывода
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def _parse_diameter(val) -> tuple[float | None, float | None]:
    """
    Извлекает диаметр из строки. В первую очередь — просто число, затем — число в [] отдельно.
    Возвращает (diameter, body_diameter): основной диаметр и диаметр тела из [].
    Примеры: 3.5 → (3.5, None); 4.5 [3.8] → (4.5, 3.8); [3.5] → (3.5, None).
    """
    if val is None:
        return (None, None)
    s = str(val).strip()
    if not s:
        return (None, None)
    s_norm = s.replace(',', '.')
    # 1. Сначала ищем просто число (до первой [)
    part_before_bracket = s_norm.split('[')[0].strip()
    diameter = None
    if part_before_bracket:
        match = re.search(r'([\d\.]+)', part_before_bracket)
        if match:
            try:
                diameter = float(match.group(1))
            except (ValueError, TypeError):
                pass
    # 2. Затем ищем число в []
    body_diameter = None
    bracket_match = re.search(r'\[([\d\.]+)\]', s_norm)
    if bracket_match:
        try:
            body_diameter = float(bracket_match.group(1))
        except (ValueError, TypeError):
            pass
    # Если простого числа нет, но есть в [] — используем его как diameter
    if diameter is None and body_diameter is not None:
        diameter = body_diameter
    return (diameter, body_diameter)


def load_sheet(ws, category: str):
    """Загружает данные из одного листа Excel."""
    print(f"  Processing sheet '{ws.title}' as category '{category}'...")
    
    # Определяем колонки по заголовкам
    header_row = 1
    # Сохраняем оригинальные заголовки для логирования
    headers_original = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row]]
    headers = [h.lower() for h in headers_original]
    
    # Находим индексы колонок
    col_indices = {
        "category": None,  # Подкатегория для протетики/лаборатории
        "line": None,
        "name": None,
        "sku": None,
        "product_type": None,
        "diameter": None,
        "length": None,
        "height": None,
        "unit": None,
        "quantity": None,
        "show_immediately": None,  # Показывать сразу или в дополнительной кнопке
    }
    
    try:
        from catalog_config import EXCEL_COLUMN_KEYWORDS, TYPE_ANGLES, DIAMETER_RANGE
        keywords_map = EXCEL_COLUMN_KEYWORDS
    except ImportError:
        keywords_map = {
            "category": ["категория", "category", "подкатегория", "subcategory", "вид", "тип товара", "кат", "cat"],
            "line": ["линейка", "line", "серия", "series"],
            "name": ["название", "name", "товар", "product", "наименование"],
            "sku": ["артикул", "sku", "код", "code", "1c", "инвойс", "invoice"],
            "product_type": ["тип", "type", "прямой", "угловой", "прямой/угловой"],
            "diameter": ["диаметр", "diameter", "диам", "ø", "d"],
            "length": ["длина", "length", "l", "высота десны", "gum height"],
            "height": ["высота абатмента", "abutment height", "высота", "height", "абатмент"],
            "unit": ["ед", "unit", "единица", "измерения", "изм"],
            "quantity": ["количество", "quantity", "кол-во", "остаток", "склад", "qty"],
            "show_immediately": ["показывать", "видимость", "visibility", "show"],
        }
        TYPE_ANGLES = (0, 17, 25, 30, 15, 20)
        DIAMETER_RANGE = (2.0, 10.0)
    
    # Для Наборы и материалы — не ищем колонки тип, диаметр, длина, высота абатмента
    skip_cols_for_no_size = ("product_type", "diameter", "length", "height")
    if category in ("Наборы", "материалы"):
        for c in skip_cols_for_no_size:
            col_indices[c] = None
        print(f"    INFO: For '{category}' — columns type, diameter, length, height are skipped (no_size items).")
    
    for col_type, keywords in keywords_map.items():
        if category in ("Наборы", "материалы") and col_type in skip_cols_for_no_size:
            continue
        for idx, header in enumerate(headers):
            # Заголовки уже в нижнем регистре, keywords тоже приводим к нижнему
            # Проверяем точное совпадение или вхождение ключевого слова
            header_lower = header.lower()
            for kw in keywords:
                kw_lower = kw.lower()
                # Проверяем точное совпадение или вхождение
                # Не используем "header in keyword" — это даёт ложные совпадения (например, "Тип" -> "тип товара").
                if header_lower == kw_lower or kw_lower in header_lower:
                    col_indices[col_type] = idx
                    print(f"    Found column '{col_type}' at index {idx}: '{headers_original[idx]}' (matched '{kw}')")
                    break
            if col_indices[col_type] is not None:
                break
    
    # Предупреждение, если колонка "Категория" не найдена (но не критично)
    if col_indices["category"] is None:
        print(f"    INFO: Category column not found. Will use standard structure without subcategories.")
        print(f"    Available headers: {', '.join(headers_original)}")
    else:
        print(f"    INFO: Category column found at index {col_indices['category']}: '{headers_original[col_indices['category']]}'")
    
    # Проверяем обязательные колонки (только название обязательно, артикул опционален)
    required = ["name"]
    missing = [col for col in required if col_indices[col] is None]
    if missing:
        print(f"    WARNING: Missing required columns: {', '.join(missing)}")
        return None
    
    # Предупреждение, если колонка с артикулом не найдена (но не критично)
    if col_indices["sku"] is None:
        print(f"    WARNING: SKU column not found. Will auto-generate SKUs from product names.")
    
    # Загружаем данные
    catalog = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    no_size_catalog = defaultdict(lambda: defaultdict(list))
    
    # Структура для хранения информации о видимости элементов
    # visibility[category][level][item_name] = show_immediately (bool)
    # level может быть: "subcategory", "line", "product"
    visibility = defaultdict(lambda: defaultdict(dict))
    
    errors = []
    loaded_count = 0
    skipped_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row or not any(row):
            continue
        
        # Извлекаем значения
        subcategory = ""
        if col_indices["category"] is not None and col_indices["category"] < len(row):
            category_val = row[col_indices["category"]]
            if category_val is not None and str(category_val).strip():
                subcategory = str(category_val).strip()
        
        line = str(row[col_indices["line"]]).strip() if col_indices["line"] is not None and col_indices["line"] < len(row) and row[col_indices["line"]] else ""
        name = str(row[col_indices["name"]]).strip() if col_indices["name"] is not None and col_indices["name"] < len(row) and row[col_indices["name"]] else ""
        sku = str(row[col_indices["sku"]]).strip() if col_indices["sku"] is not None and col_indices["sku"] < len(row) and row[col_indices["sku"]] else ""
        
        # Определяем, показывать ли элемент сразу (по умолчанию True, если колонка не указана)
        show_immediately = True
        if col_indices["show_immediately"] is not None and col_indices["show_immediately"] < len(row):
            show_val = row[col_indices["show_immediately"]]
            if show_val is not None:
                show_str = str(show_val).strip().lower()
                # Если значение "0", "нет", "no", "дополнительно", "additional" - показывать в дополнительной кнопке
                if show_str in ["0", "нет", "no", "дополнительно", "additional", "false", "скрыть", "hide"]:
                    show_immediately = False
                # Если значение "1", "да", "yes", "показывать", "show" - показывать сразу
                elif show_str in ["1", "да", "yes", "показывать", "show", "true", "видно", "visible"]:
                    show_immediately = True
        
        # Логируем для отладки (только первые несколько строк)
        if row_idx <= header_row + 5:
            if subcategory:
                print(f"    Row {row_idx}: Subcategory='{subcategory}', Line='{line}', Name='{name[:30] if len(name) > 30 else name}...'")
            elif col_indices["category"] is not None:
                print(f"    Row {row_idx}: Category column exists but is empty")
        
        # Проверяем обязательные поля (название обязательно)
        if not name:
            skipped_count += 1
            errors.append(f"    Row {row_idx}: Missing name")
            continue
        
        # Получаем тип (для протетики): градусы (0, 17, 25, 30) или предписание [N] (без шестигранника)
        # Используем TYPE_ANGLES и DIAMETER_RANGE из catalog_config: числа в диапазоне диаметров,
        # но не в списке углов — это диаметры, а не типы (избегаем 4.5, 5.5 как "типы").
        product_type = None
        if col_indices["product_type"] is not None and col_indices["product_type"] < len(row):
            type_val = row[col_indices["product_type"]]
            if type_val is not None and str(type_val).strip():
                type_str = str(type_val).strip()
                if "[n]" in type_str.lower():
                    # Предписание N (без шестигранника): "0 [N]", "17 [N]", "30 [N]" или "[N]"
                    type_str_clean = type_str.replace("[N]", "").replace("[n]", "").strip()
                    if type_str_clean and type_str_clean.replace(".", "").isdigit():
                        product_type = f"{int(float(type_str_clean))} [N]"
                    else:
                        product_type = "[N]"
                else:
                    try:
                        num_val = float(type_str)
                        int_val = int(num_val) if num_val == int(num_val) else None
                        # Число в диапазоне диаметров, но не в списке углов — это диаметр, не тип
                        if DIAMETER_RANGE[0] <= num_val <= DIAMETER_RANGE[1] and (int_val is None or int_val not in TYPE_ANGLES):
                            product_type = None  # Не тип, а диаметр (4.5, 5.5 и т.д.)
                        elif int_val is not None and int_val in TYPE_ANGLES:
                            product_type = int_val
                        elif num_val in TYPE_ANGLES:
                            product_type = num_val
                        else:
                            product_type = None  # Неизвестное значение — не используем как тип
                    except (ValueError, TypeError):
                        type_str_lower = type_str.lower()
                        if "прямой" in type_str_lower or "straight" in type_str_lower:
                            product_type = 0
                        elif "угловой" in type_str_lower or "angled" in type_str_lower or "угол" in type_str_lower:
                            product_type = 17
        
        # Получаем диаметр и длину/высоту десны
        diameter = None
        diameter_body = None
        length = None
        height = None
        
        if col_indices["diameter"] is not None and col_indices["diameter"] < len(row):
            diam_val = row[col_indices["diameter"]]
            diameter, diameter_body = _parse_diameter(diam_val)
        
        if col_indices["length"] is not None and col_indices["length"] < len(row):
            length_val = row[col_indices["length"]]
            if length_val and str(length_val).strip():
                try:
                    length = float(length_val)
                except (ValueError, TypeError):
                    pass
        
        if col_indices["height"] is not None and col_indices["height"] < len(row):
            height_val = row[col_indices["height"]]
            if height_val and str(height_val).strip():
                try:
                    height = float(height_val)
                except (ValueError, TypeError):
                    pass
        
        # Единица измерения
        unit = "шт"
        if col_indices["unit"] is not None and col_indices["unit"] < len(row):
            unit_val = row[col_indices["unit"]]
            if unit_val and str(unit_val).strip():
                unit = str(unit_val).strip()
        
        # Количество (остаток на складе)
        qty = 0
        if col_indices["quantity"] is not None and col_indices["quantity"] < len(row):
            qty_val = row[col_indices["quantity"]]
            if qty_val is not None and str(qty_val).strip():
                try:
                    qty = int(float(qty_val))
                    qty = max(0, qty)
                except (ValueError, TypeError):
                    pass
        
        # Если артикул отсутствует, генерируем его автоматически на основе названия и параметров
        if not sku:
            # Генерируем артикул на основе названия товара и параметров
            sku_parts = []
            if name:
                # Берем первые буквы из названия (до 10 символов, только буквы и цифры)
                name_short = ''.join(c for c in name[:10] if c.isalnum() or c in ['-', '_']).upper()
                if name_short:
                    sku_parts.append(name_short)
            
            # Добавляем параметры, если есть
            if diameter is not None:
                sku_parts.append(f"D{int(diameter * 10)}")
            if length is not None:
                sku_parts.append(f"L{int(length * 10)}")
            if height is not None:
                sku_parts.append(f"H{int(height * 10)}")
            if product_type is not None:
                # product_type теперь может быть числом (градусы) или строкой (для обратной совместимости)
                if isinstance(product_type, (int, float)):
                    # Для градусов используем формат типа "0", "17", "25"
                    sku_parts.append(f"{int(product_type)}")
                else:
                    # Для строк (старый формат) берем первые 2 символа
                    sku_parts.append(str(product_type)[:2].upper())
            
            sku = "-".join(sku_parts) if sku_parts else f"AUTO-{row_idx}"
            # Логируем только первые несколько случаев для отладки
            if loaded_count < 5:
                print(f"    Row {row_idx}: Auto-generated SKU: {sku}")
        
        # Формируем объект товара
        product_data = {
            "name": name,
            "sku": sku,
            "unit": unit,
            "qty": qty,
            "show_immediately": show_immediately,  # Информация о видимости из Excel
        }
        if diameter_body is not None:
            product_data["diameter_body"] = diameter_body  # диаметр тела из []
        
        # Добавляем в каталог в зависимости от категории
        # Протетика/Лаборатория/Наборы/материалы: Category = колонка "Категория", Sub_category = колонка "Линейка"
        # Структура: category -> Category -> Sub_category -> product -> type -> diameter -> length -> height
        # Название товара используется как у имплантов (в product_data)
        if category in ["Протетика", "Лаборатория"] and product_type and diameter is not None and length is not None:
            product_key = name
            excel_cat = subcategory.strip() if subcategory else "Основное"  # Category
            excel_line = line.strip() if line else "Без линейки"           # Sub_category
            
            # Видимость: Category (subcategory), Sub_category (line)
            if excel_cat not in visibility[category]["subcategory"]:
                visibility[category]["subcategory"][excel_cat] = show_immediately
            else:
                visibility[category]["subcategory"][excel_cat] = visibility[category]["subcategory"][excel_cat] or show_immediately
            if excel_line not in visibility[category]["line"]:
                visibility[category]["line"][excel_line] = show_immediately
            else:
                visibility[category]["line"][excel_line] = visibility[category]["line"][excel_line] or show_immediately
            if product_key not in visibility[category]["product"]:
                visibility[category]["product"][product_key] = show_immediately
            else:
                visibility[category]["product"][product_key] = visibility[category]["product"][product_key] or show_immediately
            
            if excel_cat not in catalog[category]:
                catalog[category][excel_cat] = {}
            if excel_line not in catalog[category][excel_cat]:
                catalog[category][excel_cat][excel_line] = {}
            if product_key not in catalog[category][excel_cat][excel_line]:
                catalog[category][excel_cat][excel_line][product_key] = {}
            if product_type not in catalog[category][excel_cat][excel_line][product_key]:
                catalog[category][excel_cat][excel_line][product_key][product_type] = {}
            if diameter not in catalog[category][excel_cat][excel_line][product_key][product_type]:
                catalog[category][excel_cat][excel_line][product_key][product_type][diameter] = {}
            if length not in catalog[category][excel_cat][excel_line][product_key][product_type][diameter]:
                catalog[category][excel_cat][excel_line][product_key][product_type][diameter][length] = {}
            
            if height is not None:
                catalog[category][excel_cat][excel_line][product_key][product_type][diameter][length][height] = product_data
            else:
                catalog[category][excel_cat][excel_line][product_key][product_type][diameter][length] = product_data
            loaded_count += 1
        elif category in ["Протетика", "Лаборатория"] and diameter is not None and length is not None:
            # Протетика/Лаборатория без типа (но с размерами): Category -> Sub_category -> product -> diameter -> length -> height
            product_key = name
            excel_cat = subcategory.strip() if subcategory else "Основное"
            excel_line = line.strip() if line else "Без линейки"
            
            if excel_cat not in visibility[category]["subcategory"]:
                visibility[category]["subcategory"][excel_cat] = show_immediately
            else:
                visibility[category]["subcategory"][excel_cat] = visibility[category]["subcategory"][excel_cat] or show_immediately
            if excel_line not in visibility[category]["line"]:
                visibility[category]["line"][excel_line] = show_immediately
            else:
                visibility[category]["line"][excel_line] = visibility[category]["line"][excel_line] or show_immediately
            if product_key not in visibility[category]["product"]:
                visibility[category]["product"][product_key] = show_immediately
            else:
                visibility[category]["product"][product_key] = visibility[category]["product"][product_key] or show_immediately
            
            if excel_cat not in catalog[category]:
                catalog[category][excel_cat] = {}
            if excel_line not in catalog[category][excel_cat]:
                catalog[category][excel_cat][excel_line] = {}
            if product_key not in catalog[category][excel_cat][excel_line]:
                catalog[category][excel_cat][excel_line][product_key] = {}
            if diameter not in catalog[category][excel_cat][excel_line][product_key]:
                catalog[category][excel_cat][excel_line][product_key][diameter] = {}
            if length not in catalog[category][excel_cat][excel_line][product_key][diameter]:
                catalog[category][excel_cat][excel_line][product_key][diameter][length] = {}
            
            if height is not None:
                catalog[category][excel_cat][excel_line][product_key][diameter][length][height] = product_data
            else:
                catalog[category][excel_cat][excel_line][product_key][diameter][length] = product_data
            loaded_count += 1
        elif diameter is not None and length is not None:
            # Импланты: category -> line -> diameter_key -> length
            # diameter_key: float для "4.5", tuple (4.5, 3.8) для "4.5 [3.8]" — это разные продукты
            if not line:
                line = "Без линейки"
            # Для имплантов используем специальную структуру без product
            if category == "Импланты":
                if line not in catalog[category]:
                    catalog[category][line] = {}
                    # Сохраняем информацию о видимости линейки
                    visibility[category]["line"][line] = show_immediately
                else:
                    # Если линейка уже есть, обновляем видимость (True, если хотя бы один товар имеет True)
                    if line not in visibility[category]["line"]:
                        visibility[category]["line"][line] = show_immediately
                    else:
                        visibility[category]["line"][line] = visibility[category]["line"][line] or show_immediately
                # Ключ диаметра: (diameter, diameter_body) если есть [], иначе float diameter
                diam_key = (diameter, diameter_body) if diameter_body is not None else diameter
                if diam_key not in catalog[category][line]:
                    catalog[category][line][diam_key] = {}
                catalog[category][line][diam_key][length] = product_data
            else:
                # Для других категорий с размерами (если есть)
                product_key = name
                if product_key not in catalog[category]:
                    catalog[category][product_key] = {}
                    # Сохраняем информацию о видимости товара
                    visibility[category]["product"][product_key] = show_immediately
                if line not in catalog[category][product_key]:
                    catalog[category][product_key][line] = {}
                    # Сохраняем информацию о видимости линейки
                    visibility[category]["line"][line] = show_immediately
                if diameter not in catalog[category][product_key][line]:
                    catalog[category][product_key][line][diameter] = {}
                catalog[category][product_key][line][diameter][length] = product_data
            loaded_count += 1
        else:
            # Товар без размеров
            if not line:
                line = "Без категории"
            if category == "Импланты":
                # Для имплантов без размеров сохраняем в catalog
                if line not in catalog[category]:
                    catalog[category][line] = {}
                    visibility[category]["line"][line] = show_immediately
                else:
                    if line not in visibility[category]["line"]:
                        visibility[category]["line"][line] = show_immediately
                    else:
                        visibility[category]["line"][line] = visibility[category]["line"][line] or show_immediately
                if "no_size" not in catalog[category][line]:
                    catalog[category][line]["no_size"] = []
                catalog[category][line]["no_size"].append(product_data)
            elif category in ["Протетика", "Лаборатория", "Наборы", "материалы"]:
                # Category -> Sub_category -> no_size list
                excel_cat = subcategory.strip() if subcategory else "Основное"
                excel_line = line.strip() if line else "Без линейки"
                if excel_cat not in visibility[category]["subcategory"]:
                    visibility[category]["subcategory"][excel_cat] = show_immediately
                else:
                    visibility[category]["subcategory"][excel_cat] = visibility[category]["subcategory"][excel_cat] or show_immediately
                if excel_line not in visibility[category]["line"]:
                    visibility[category]["line"][excel_line] = show_immediately
                else:
                    visibility[category]["line"][excel_line] = visibility[category]["line"][excel_line] or show_immediately
                if excel_cat not in catalog[category]:
                    catalog[category][excel_cat] = {}
                if excel_line not in catalog[category][excel_cat]:
                    catalog[category][excel_cat][excel_line] = {}
                if "no_size" not in catalog[category][excel_cat][excel_line]:
                    catalog[category][excel_cat][excel_line]["no_size"] = []
                catalog[category][excel_cat][excel_line]["no_size"].append(product_data)
            else:
                no_size_catalog[category][line].append(product_data)
            loaded_count += 1
    
    if errors:
        print(f"    Warnings ({len(errors)} rows skipped):")
        for error in errors[:5]:
            print(error)
        if len(errors) > 5:
            print(f"    ... and {len(errors) - 5} more")
    
    print(f"    Loaded: {loaded_count} items, Skipped: {skipped_count} rows")
    return catalog, no_size_catalog, dict(visibility)

def load_catalog_from_excel(filename: str = "catalog_template.xlsx"):
    """Загружает каталог из Excel файла с отдельными листами для каждой категории."""
    file_path = Path(filename)
    if not file_path.exists():
        print(f"ERROR: File {filename} not found!")
        print("Available files:")
        for xlsx_file in Path(".").glob("*.xlsx"):
            print(f"  - {xlsx_file}")
        return None
    
    wb = load_workbook(filename, data_only=True)
    
    print(f"Loading: {filename}")
    print(f"Sheets: {wb.sheetnames}\n")
    
    try:
        from catalog_config import SHEET_TO_CATEGORY, SKIP_SHEETS
    except ImportError:
        SHEET_TO_CATEGORY = {
            "импланты": "Импланты", "implants": "Импланты",
            "протетика": "Протетика", "prosthetics": "Протетика",
            "лаборатория": "Лаборатория", "laboratory": "Лаборатория", "lab": "Лаборатория",
            "наборы": "Наборы", "kits": "Наборы",
            "материалы": "материалы", "materials": "материалы",
        }
        SKIP_SHEETS = ["инструкция", "instruction", "readme"]
    
    all_catalog = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    all_no_size = defaultdict(lambda: defaultdict(list))
    all_visibility = defaultdict(lambda: defaultdict(dict))  # Собираем информацию о видимости
    
    # Обрабатываем каждый лист
    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(skip in sheet_lower for skip in SKIP_SHEETS):
            continue

        # Определяем категорию по названию листа
        category = SHEET_TO_CATEGORY.get(sheet_lower)
        if not category:
            for key, cat in SHEET_TO_CATEGORY.items():
                if key in sheet_lower:
                    category = cat
                    break
        if not category:
            category = sheet_name
        
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} columns)")
        
        result = load_sheet(ws, category)
        if result:
            sheet_catalog, sheet_no_size, sheet_visibility = result
            
            # Объединяем информацию о видимости
            for cat, levels in sheet_visibility.items():
                if cat not in all_visibility:
                    all_visibility[cat] = defaultdict(dict)
                for level, items in levels.items():
                    for item_name, show_immediately in items.items():
                        # Если элемент уже есть, берем True (показывать сразу) если хотя бы один товар имеет True
                        if item_name not in all_visibility[cat][level]:
                            all_visibility[cat][level][item_name] = show_immediately
                        else:
                            # Если хотя бы один товар должен показываться сразу, элемент показывается сразу
                            all_visibility[cat][level][item_name] = all_visibility[cat][level][item_name] or show_immediately
            
            # Объединяем с общим каталогом (рекурсивно: поддерживаем Category -> Sub_category -> product для Протетики/Лаборатории)
            def merge_catalog_dicts(d1, d2):
                """Рекурсивно объединяет каталог: словари мержатся, списки (no_size) объединяются."""
                for key, value in d2.items():
                    if key not in d1:
                        d1[key] = value.copy() if isinstance(value, list) else value
                    elif isinstance(d1[key], dict) and isinstance(value, dict):
                        merge_catalog_dicts(d1[key], value)
                    elif isinstance(d1[key], list) and isinstance(value, list):
                        d1[key].extend(value)
                    else:
                        d1[key] = value.copy() if isinstance(value, list) else value
            
            for cat, sheet_data in sheet_catalog.items():
                if cat not in all_catalog:
                    all_catalog[cat] = {}
                merge_catalog_dicts(all_catalog[cat], sheet_data)
            
            for cat, lines in sheet_no_size.items():
                for line, products in lines.items():
                    if cat not in all_no_size:
                        all_no_size[cat] = {}
                    if line not in all_no_size[cat]:
                        all_no_size[cat][line] = []
                    all_no_size[cat][line].extend(products)
    
    # Вспомогательная функция для обработки line_data
    def _process_prosthetics_line_data(result, cat, subcategory, product, line, line_data, has_subcategory=False):
        """Обрабатывает line_data для протетики/лаборатории с учетом подкатегорий"""
        if has_subcategory:
            target = result[cat][subcategory][product][line]
        else:
            target = result[cat][product][line]
        
        # Проверяем, это протетика/лаборатория с типами или стандартная структура
        if cat in ["Протетика", "Лаборатория"] and any(isinstance(k, (int, float)) for k in line_data.keys()):
            # Протетика/Лаборатория с типами (градусы)
            for product_type, type_data in line_data.items():
                target[product_type] = {}
                for diam, diameters_data in type_data.items():
                    target[product_type][diam] = {}
                    for length, lengths_data in diameters_data.items():
                        if isinstance(lengths_data, dict) and any(isinstance(k, (int, float)) for k in lengths_data.keys()):
                            # Есть высоты абатмента
                            target[product_type][diam][length] = {}
                            for height, product_info in lengths_data.items():
                                if isinstance(height, (int, float)):
                                    target[product_type][diam][length][height] = product_info
                        else:
                            # Только высота десны
                            target[product_type][diam][length] = lengths_data
        elif cat in ["Протетика", "Лаборатория"]:
            # Протетика/Лаборатория без типа
            for diam, diameters_data in line_data.items():
                if isinstance(diam, (int, float)):
                    target[diam] = {}
                    for length, lengths_data in diameters_data.items():
                        if isinstance(lengths_data, dict):
                            if any(isinstance(k, (int, float)) for k in lengths_data.keys()):
                                # Есть высоты абатмента
                                target[diam][length] = {}
                                for height, product_info in lengths_data.items():
                                    if isinstance(height, (int, float)) and isinstance(product_info, dict) and "name" in product_info:
                                        target[diam][length][height] = product_info
                            elif isinstance(lengths_data, dict) and "name" in lengths_data:
                                # Только высота десны (данные товара)
                                target[diam][length] = lengths_data
                        elif isinstance(lengths_data, dict) and "name" in lengths_data:
                            # Просто данные товара
                            target[diam][length] = lengths_data
        else:
            # Другая структура
            target.update(line_data)
    
    # Объединяем данные
    result = {}
    for cat, data in all_catalog.items():
        result[cat] = {}
        
        if cat == "Импланты":
            # Для имплантов: category -> line -> diameter -> length (или no_size -> list)
            for line, line_data in data.items():
                result[cat][line] = {}
                for diam, diameters_data in line_data.items():
                    if diam == "no_size":
                        result[cat][line]["no_size"] = diameters_data
                        continue
                    result[cat][line][diam] = {}
                    for length, product_info in diameters_data.items():
                        result[cat][line][diam][length] = product_info
        elif cat in ["Протетика", "Лаборатория", "Наборы", "материалы"]:
            # Category (excel_cat) -> Sub_category (excel_line) -> product -> type -> diameter -> length -> height (или no_size)
            for excel_cat, excel_line_data in data.items():
                result[cat][excel_cat] = {}
                for excel_line, product_data in excel_line_data.items():
                    result[cat][excel_cat][excel_line] = {}
                    for product_key, line_data in product_data.items():
                        if product_key == "no_size":
                            result[cat][excel_cat][excel_line]["no_size"] = line_data
                            continue
                        result[cat][excel_cat][excel_line][product_key] = {}
                        # line_data здесь: product_type -> diameter -> length -> height или diameter -> length -> height
                        if isinstance(line_data, dict) and any(isinstance(k, (int, float)) for k in line_data.keys()):
                            for product_type, type_data in line_data.items():
                                if not isinstance(type_data, dict):
                                    continue
                                result[cat][excel_cat][excel_line][product_key][product_type] = {}
                                for diam, diameters_data in type_data.items():
                                    if not isinstance(diam, (int, float)):
                                        continue
                                    result[cat][excel_cat][excel_line][product_key][product_type][diam] = {}
                                    for length, lengths_data in diameters_data.items():
                                        if isinstance(lengths_data, dict) and "name" in lengths_data:
                                            result[cat][excel_cat][excel_line][product_key][product_type][diam][length] = lengths_data
                                        elif isinstance(lengths_data, dict):
                                            for height, product_info in lengths_data.items():
                                                if isinstance(height, (int, float)) and isinstance(product_info, dict) and "name" in product_info:
                                                    if length not in result[cat][excel_cat][excel_line][product_key][product_type][diam]:
                                                        result[cat][excel_cat][excel_line][product_key][product_type][diam][length] = {}
                                                    result[cat][excel_cat][excel_line][product_key][product_type][diam][length][height] = product_info
                        elif isinstance(line_data, dict):
                            # Без типа: diameter -> length -> height
                            for diam, diameters_data in line_data.items():
                                if not isinstance(diam, (int, float)):
                                    continue
                                result[cat][excel_cat][excel_line][product_key][diam] = {}
                                for length, lengths_data in diameters_data.items():
                                    if isinstance(lengths_data, dict) and "name" in lengths_data:
                                        result[cat][excel_cat][excel_line][product_key][diam][length] = lengths_data
                                    elif isinstance(lengths_data, dict):
                                        for height, product_info in lengths_data.items():
                                            if isinstance(height, (int, float)) and isinstance(product_info, dict) and "name" in product_info:
                                                if length not in result[cat][excel_cat][excel_line][product_key][diam]:
                                                    result[cat][excel_cat][excel_line][product_key][diam][length] = {}
                                                result[cat][excel_cat][excel_line][product_key][diam][length][height] = product_info
        else:
            # Остальные категории (старая структура): product -> line -> ...
            first_level_keys = list(data.keys())
            has_subcategories = False
            if first_level_keys:
                first_item = data[first_level_keys[0]]
                if isinstance(first_item, dict):
                    inner_keys = list(first_item.keys()) if first_item else []
                    if inner_keys:
                        inner_first = first_item[inner_keys[0]]
                        if isinstance(inner_first, dict):
                            has_subcategories = True
            if has_subcategories:
                for subcategory, subcategory_data in data.items():
                    result[cat][subcategory] = {}
                    for product, product_data in subcategory_data.items():
                        result[cat][subcategory][product] = {}
                        for line, line_data in product_data.items():
                            result[cat][subcategory][product][line] = {}
                            _process_prosthetics_line_data(result, cat, subcategory, product, line, line_data, has_subcategory=True)
            else:
                for product, product_data in data.items():
                    result[cat][product] = {}
                    for line, line_data in product_data.items():
                        result[cat][product][line] = {}
                        _process_prosthetics_line_data(result, cat, None, product, line, line_data, has_subcategory=False)
    
    # Добавляем товары без размеров (Протетика/Лаборатория/Наборы/материалы уже в result через catalog)
    for cat, lines in all_no_size.items():
        if cat not in result:
            result[cat] = {}
        if cat in ["Протетика", "Лаборатория", "Наборы", "материалы"]:
            continue
        for line, products in lines.items():
            if cat == "Импланты":
                if line not in result[cat]:
                    result[cat][line] = {}
                result[cat][line]["no_size"] = products
            else:
                if products:
                    product_key = products[0]["name"] if products else line
                    if product_key not in result[cat]:
                        result[cat][product_key] = {}
                    if line not in result[cat][product_key]:
                        result[cat][product_key][line] = {}
                    result[cat][product_key][line]["no_size"] = products
                else:
                    if line not in result[cat]:
                        result[cat][line] = {}
                    result[cat][line]["no_size"] = []
    
    return result, dict(all_visibility)


def _product_line(product_info) -> str:
    """Строка для вывода product_info (name, sku, unit, qty, diameter_body) в catalog_data.py."""
    def esc(s: str) -> str:
        return str(s).replace("\\", "\\\\").replace('"', '\\"')
    if not isinstance(product_info, dict):
        return '"name": "", "sku": "", "unit": "шт", "qty": 0'
    parts = [
        f'"name": "{esc(product_info.get("name", ""))}", ',
        f'"sku": "{esc(product_info.get("sku", ""))}", ',
        f'"unit": "{esc(product_info.get("unit", "шт"))}", ',
        f'"qty": {product_info.get("qty", 0)}',
    ]
    if product_info.get("diameter_body") is not None:
        parts.append(f', "diameter_body": {product_info["diameter_body"]}')
    return "".join(parts)


def _generate_line_data(lines, cat, line, line_data, indent=12):
    """Генерирует код для line_data с учетом структуры протетики/лаборатории"""
    indent_str = " " * indent
    if isinstance(line_data, dict) and any(isinstance(k, (int, float)) for k in line_data.keys()):
        # Протетика/Лаборатория с типами (градусы)
        lines.append(f'{indent_str}"{line}": {{')  # Линейка импланта
        for product_type, type_data in _sort_mixed_keys(line_data.items()):
            lines.append(f'{indent_str}    "{product_type}": {{')
            for diam, diameters_data in _sort_mixed_keys(type_data.items()):
                lines.append(f"{indent_str}        {diam}: {{")
                for length, lengths_data in _sort_mixed_keys(diameters_data.items()):
                    if isinstance(lengths_data, dict) and any(isinstance(k, (int, float)) for k in lengths_data.keys()):
                        # Есть высоты абатмента
                        lines.append(f"{indent_str}            {length}: {{")
                        for height, product_info in _sort_mixed_keys(lengths_data.items()):
                            if isinstance(height, (int, float)):
                                lines.append(f"{indent_str}                {height}: {{{_product_line(product_info)} }},")
                        lines.append(f"{indent_str}            }},")
                    else:
                        # Только высота десны
                        if isinstance(lengths_data, dict) and "name" in lengths_data:
                            lines.append(f"{indent_str}            {length}: {{{_product_line(lengths_data)} }},")
                        else:
                            logging.getLogger(__name__).warning(
                                "Пропуск невалидного product_info для линии %s (length=%s): не dict или нет 'name'",
                                line, length
                            )
                lines.append(f"{indent_str}        }},")
            lines.append(f"{indent_str}    }},")
        lines.append(f"{indent_str}}},")
    elif isinstance(line_data, dict) and any(isinstance(k, (int, float)) for k in line_data.keys()):
        # Стандартная структура с диаметрами (для протетики/лаборатории без типа)
        lines.append(f'{indent_str}"{line}": {{')
        for diam, diameters_data in _sort_mixed_keys(line_data.items()):
            if isinstance(diam, (int, float)):
                lines.append(f"{indent_str}        {diam}: {{")
                for length, lengths_data in _sort_mixed_keys(diameters_data.items()):
                    if isinstance(lengths_data, dict):
                        if any(isinstance(k, (int, float)) for k in lengths_data.keys()):
                            # Есть высоты абатмента
                            lines.append(f"{indent_str}            {length}: {{")
                            for height, product_info in _sort_mixed_keys(lengths_data.items()):
                                if isinstance(height, (int, float)) and isinstance(product_info, dict) and "name" in product_info:
                                    lines.append(f"{indent_str}                {height}: {{{_product_line(product_info)} }},")
                            lines.append(f"{indent_str}            }},")
                        elif "name" in lengths_data:
                            # Только высота десны
                            lines.append(f"{indent_str}            {length}: {{{_product_line(lengths_data)} }},")
                    elif isinstance(lengths_data, dict) and "name" in lengths_data:
                        lines.append(f"{indent_str}            {length}: {{{_product_line(lengths_data)} }},")
                lines.append(f"{indent_str}        }},")
        lines.append(f"{indent_str}}},")
    elif "no_size" in line_data:
        # Товары без размеров
        lines.append(f'{indent_str}"{line}": {{')
        lines.append(f'{indent_str}    "no_size": [')
        for item in line_data["no_size"]:
            lines.append(f'{indent_str}        {{{_product_line(item)} }},')
        lines.append(f'{indent_str}    ],')
        lines.append(f"{indent_str}}},")
    else:
        # Другая структура
        lines.append(f'{indent_str}"{line}": {_format_dict(line_data, indent + 4)},')

def _format_dict(d, indent=0):
    """Форматирует словарь для вывода в Python код"""
    if not isinstance(d, dict):
        return repr(d)
    indent_str = " " * indent
    lines = ["{"]
    for k, v in sorted(d.items()):
        if isinstance(v, dict):
            lines.append(f'{indent_str}    "{k}": {_format_dict(v, indent + 4)},')
        else:
            lines.append(f'{indent_str}    "{k}": {repr(v)},')
    lines.append(f"{indent_str}}}")
    return "\n".join(lines)

def _sort_mixed_keys(items):
    """Сортирует (k, v): сначала числа, затем tuple (d, body), затем 'no_size', затем остальные строки."""
    def key(pair):
        k = pair[0]
        if isinstance(k, (int, float)):
            return (0, k, 0)
        if isinstance(k, tuple) and len(k) == 2:
            return (0, k[0], k[1] if k[1] is not None else 0)
        if k == "no_size":
            return (2, 0, 0)
        return (1, str(k))
    return sorted(items, key=key)


def generate_catalog_file(catalog: dict, visibility: dict = None):
    """Генерирует catalog_data.py из загруженных данных."""
    lines = [
        '"""',
        "Данные каталога продукции (загружено из Excel).",
        "Не редактировать вручную! Используйте load_catalog_from_excel.py",
        "",
        "Структура:",
        "  CATALOG[category][product][line][type][diameter][gum_height][abutment_height] = {",
        "    'name': 'Название товара',",
        "    'sku': 'Артикул из 1C',",
        "    'unit': 'Ед. изм.'",
        "  }",
        "",
        "  CATALOG[category][line]['no_size'] = [",
        "    {'name': '...', 'sku': '...', 'unit': '...'},",
        "    ...",
        "  ]",
        "",
        "VISIBILITY[category][level][item_name] = True/False",
        "  level может быть: 'subcategory', 'line', 'product'",
        "  True = показывать сразу, False = показывать в дополнительной кнопке",
        '"""',
        "",
        "CATALOG = {",
    ]
    
    for cat, data_dict in sorted(catalog.items()):
        lines.append(f'    "{cat}": {{')
        
        if cat == "Импланты":
            # Для имплантов: category -> line -> diameter_key -> length (diameter_key: float или (d, body))
            for line, line_data in sorted(data_dict.items()):
                lines.append(f'        "{line}": {{')
                for diam, diameters_data in _sort_mixed_keys(line_data.items()):
                    if isinstance(diam, tuple) and len(diam) == 2:
                        lines.append(f"            {diam}: {{")
                        for length, product_info in sorted(diameters_data.items()):
                            lines.append(f"                {length}: {{{_product_line(product_info)} }},")
                        lines.append("            },")
                    elif isinstance(diam, (int, float)):
                        lines.append(f"            {diam}: {{")
                        for length, product_info in sorted(diameters_data.items()):
                            lines.append(f"                {length}: {{{_product_line(product_info)} }},")
                        lines.append("            },")
                    elif diam == "no_size":
                        lines.append('            "no_size": [')
                        for item in diameters_data:
                            lines.append(f'                {{{_product_line(item)} }},')
                        lines.append('            ],')
                lines.append("        },")
        elif cat in ["Протетика", "Лаборатория", "Наборы", "материалы"]:
            # Category (excel_cat) -> Sub_category (excel_line) -> product -> type -> diameter -> length -> height (или no_size)
            for excel_cat, excel_line_data in sorted(data_dict.items()):
                lines.append(f'        "{excel_cat}": {{')
                for excel_line, product_data in sorted(excel_line_data.items()):
                    lines.append(f'            "{excel_line}": {{')
                    for product_key, line_data in sorted(product_data.items(), key=lambda x: (x[0] == "no_size", str(x[0]))):
                        if product_key == "no_size":
                            lines.append('                "no_size": [')
                            for item in line_data:
                                lines.append(f'                    {{{_product_line(item)} }},')
                            lines.append('                ],')
                        else:
                            _generate_line_data(lines, cat, product_key, line_data, indent=20)
                    lines.append("            },")
                lines.append("        },")
        else:
            # Остальные категории (старая структура): product -> line -> ...
            first_level_keys = list(data_dict.keys())
            has_subcategories = False
            if first_level_keys:
                first_item = data_dict[first_level_keys[0]]
                if isinstance(first_item, dict):
                    inner_keys = list(first_item.keys()) if first_item else []
                    if inner_keys:
                        inner_first = first_item[inner_keys[0]]
                        if isinstance(inner_first, dict):
                            has_subcategories = True
            if has_subcategories:
                for subcategory, subcategory_data in sorted(data_dict.items()):
                    lines.append(f'        "{subcategory}": {{')
                    for product, product_data in sorted(subcategory_data.items()):
                        lines.append(f'            "{product}": {{')
                        for line, line_data in sorted(product_data.items()):
                            _generate_line_data(lines, cat, line, line_data, indent=16)
                        lines.append("            },")
                    lines.append("        },")
            else:
                for product, product_data in sorted(data_dict.items()):
                    lines.append(f'        "{product}": {{')
                    for line, line_data in sorted(product_data.items()):
                        _generate_line_data(lines, cat, line, line_data, indent=12)
                    lines.append("        },")
        lines.append("    },")
    
    lines.append("}")
    lines.append("")
    
    # Генерируем VISIBILITY
    if visibility:
        lines.append("VISIBILITY = {")
        for cat, levels in sorted(visibility.items()):
            lines.append(f'    "{cat}": {{')
            for level, items in sorted(levels.items()):
                lines.append(f'        "{level}": {{')
                for item_name, show_immediately in sorted(items.items()):
                    lines.append(f'            "{item_name}": {show_immediately},')
                lines.append("        },")
            lines.append("    },")
        lines.append("}")
    else:
        lines.append("VISIBILITY = {}")
    lines.append("")
    
    content = "\n".join(lines)
    
    with open("catalog_data.py", "w", encoding="utf-8") as f:
        f.write(content)
    
    print("\nGenerated: catalog_data.py")

def print_statistics(catalog: dict):
    """Выводит статистику по каталогу."""
    total_categories = len(catalog)
    total_items_with_size = 0
    total_items_no_size = 0
    
    for cat, products in catalog.items():
        for product, product_data in products.items():
            for key, value in product_data.items():
                if key == "no_size":
                    if isinstance(value, list):
                        total_items_no_size += len(value)
                    else:
                        total_items_no_size += 1
                elif isinstance(value, dict):
                    # Рекурсивно считаем товары с размерами
                    def count_items(data, depth=0):
                        count = 0
                        if depth > 10:  # Защита от бесконечной рекурсии
                            return 0
                        for k, v in data.items():
                            if isinstance(v, dict):
                                if "name" in v and "sku" in v:
                                    count += 1
                                else:
                                    count += count_items(v, depth + 1)
                        return count
                    total_items_with_size += count_items(value)
    
    print(f"\nStatistics:")
    print(f"  Categories: {total_categories}")
    print(f"  Items with sizes: {total_items_with_size}")
    print(f"  Items without sizes: {total_items_no_size}")
    print(f"  Total items: {total_items_with_size + total_items_no_size}")

def main():
    import sys
    
    filename = sys.argv[1] if len(sys.argv) > 1 else "catalog_template.xlsx"
    
    # Проверяем альтернативные имена
    if not Path(filename).exists():
        for alt_name in ["Megagenbot.xlsx", "catalog.xlsx"]:
            if Path(alt_name).exists():
                filename = alt_name
                print(f"Using alternative file: {filename}")
                break
    
    result = load_catalog_from_excel(filename)
    if result:
        catalog, visibility = result
        print_statistics(catalog)
        generate_catalog_file(catalog, visibility)
        
        # Также сохраняем JSON
        with open("catalog.json", "w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)
        print("Generated: catalog.json")
        
        print("\nNext step: Restart the bot to use the new catalog data.")

if __name__ == "__main__":
    main()
