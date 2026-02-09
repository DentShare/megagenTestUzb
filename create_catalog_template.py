"""
Создание точного шаблона Excel для каталога продукции с отдельными листами для каждой категории.

Структура:
- Каждый лист = одна категория (Импланты, Протетика, Лаборатория, Наборы, материалы)
- Для Протетики/Лаборатории/Наборов/материалов: Категория (опц.) | Линейка | Название | Артикул | ...
- Для Имплантов: Линейка | Название | Артикул | ...
- Настройка: catalog_config.py — CATEGORY_STRUCTURE определяет, нужна ли колонка "Категория"

Использование:
    python create_catalog_template.py
"""
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Устанавливаем UTF-8 для вывода
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def create_sheet_template(ws, category: str):
    """Создаёт шаблон для одного листа категории."""
    try:
        from catalog_config import CATEGORY_STRUCTURE
        has_subcategory = CATEGORY_STRUCTURE.get(category, {}).get("has_subcategory", False)
    except ImportError:
        has_subcategory = category in ["Протетика", "Лаборатория", "Наборы", "материалы"]

    # Заголовки: для категорий с подкатегориями добавляем колонку "Категория"
    base_headers = [
        "Линейка",
        "Название товара",
        "Артикул (1C)",
        "Тип",  # Для протетики: 0, 17, 25, 30 или прямой/угловой
        "Диаметр",
        "Длина",  # Для имплантов - длина, для протетики - высота десны
        "Высота абатмента",  # Только для протетики/лаборатории
        "Ед. изм.",
    ]
    if has_subcategory:
        headers = ["Категория"] + base_headers
    else:
        headers = base_headers
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Примеры данных в зависимости от категории
    examples = []
    
    if category == "Импланты":
        examples = [
            ["AnyRidge", "AnyRidge 3507", "AR-3507", "", 3.5, 7.0, "", "шт"],
            ["AnyRidge", "AnyRidge 3508", "AR-3508", "", 3.5, 8.0, "", "шт"],
            ["AnyRidge", "AnyRidge 3510", "AR-3510", "", 3.5, 10.0, "", "шт"],
            ["AnyRidge", "AnyRidge 4007", "AR-4007", "", 4.0, 7.0, "", "шт"],
            ["AnyOne", "AnyOne 3507 C", "AO-3507C", "", 3.5, 7.0, "", "шт"],
            ["AnyOne", "AnyOne 3510 C", "AO-3510C", "", 3.5, 10.0, "", "шт"],
        ]
    elif category == "Протетика":
        # С колонкой "Категория" (EzPost, ZrGen и т.д.)
        examples = [
            ["EzPost", "AnyRidge", "EZ Post Abutment", "EP-AR-45-15-25-ST", 0, 4.5, 1.5, 2.5, "шт"],
            ["EzPost", "AnyRidge", "EZ Post Abutment", "EP-AR-45-20-25-ST", 0, 4.5, 2.0, 2.5, "шт"],
            ["EzPost", "AnyRidge", "EZ Post Abutment", "EP-AR-45-15-25-AN", 17, 4.5, 1.5, 2.5, "шт"],
            ["EzPost", "AnyOne", "EZ Post Abutment", "EP-AO-45-15-25-ST", 0, 4.5, 1.5, 2.5, "шт"],
            ["EzPost", "AnyRidge", "Healing Abutment", "HA-AR-45-15-25-ST", 0, 4.5, 1.5, 2.5, "шт"],
        ]
    elif category == "Лаборатория":
        examples = [
            ["LabAnalog", "AnyRidge", "Lab Analog", "LA-AR-45-15-25", "", 4.5, 1.5, 2.5, "шт"],
            ["LabAnalog", "AnyOne", "Lab Analog", "LA-AO-45-15-25", "", 4.5, 1.5, 2.5, "шт"],
            ["LabAnalog", "AnyRidge", "Lab Analog (EZ) EILA 400", "LA-EZ-400", "", "", "", "", "шт"],
        ]
    elif category == "Наборы":
        examples = [
            ["Основное", "", "Anchor Kit (AO) KAGAS3001", "KAGAS3001", "", "", "", "", "компл"],
            ["Основное", "", "Anchor Kit (AR) KAGAS3000", "KAGAS3000", "", "", "", "", "компл"],
        ]
    elif category == "материалы":
        examples = [
            ["Основное", "", "Fiziodisfensor COXO", "FZ-COXO", "", "", "", "", "шт"],
            ["Основное", "", "Fiziodisfensor Ki-20", "FZ-KI20", "", "", "", "", "шт"],
        ]
    
    # Записываем примеры
    diam_cols = [5, 6, 7] if not has_subcategory else [6, 7, 8]  # Диаметр, Длина, Высота абатмента
    sku_col = 3 if not has_subcategory else 4  # Артикул
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx in diam_cols:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == sku_col:
                cell.font = Font(name="Courier New")  # Моноширинный шрифт для артикулов
    
    # Настройка ширины колонок (с учётом колонки "Категория" при has_subcategory)
    if has_subcategory:
        column_widths = {
            "A": 15,  # Категория
            "B": 20,  # Линейка
            "C": 40,  # Название товара
            "D": 20,  # Артикул (1C)
            "E": 12,  # Тип
            "F": 10,  # Диаметр
            "G": 12,  # Длина/Высота десны
            "H": 15,  # Высота абатмента
            "I": 10,  # Ед. изм.
        }
    else:
        column_widths = {
            "A": 20,  # Линейка
            "B": 40,  # Название товара
            "C": 20,  # Артикул (1C)
            "D": 12,  # Тип
            "E": 10,  # Диаметр
            "F": 12,  # Длина/Высота десны
            "G": 15,  # Высота абатмента
            "H": 10,  # Ед. изм.
        }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Заморозить первую строку
    ws.freeze_panes = "A2"

def create_catalog_template():
    """Создаёт точный Excel шаблон для каталога с отдельными листами."""
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Создаём листы для каждой категории
    categories = ["Импланты", "Протетика", "Лаборатория", "Наборы", "материалы"]
    
    for category in categories:
        ws = wb.create_sheet(category)
        create_sheet_template(ws, category)
    
    # Добавить инструкции на отдельный лист
    ws_instructions = wb.create_sheet("Инструкция", 0)
    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ КАТАЛОГА",
        "",
        "1. СТРУКТУРА ФАЙЛА:",
        "   - Каждый лист = одна категория товаров",
        "   - Листы: Импланты, Протетика, Лаборатория, Наборы, материалы",
        "   - На каждом листе заполняйте товары только этой категории",
        "",
        "2. СТРУКТУРА ТАБЛИЦЫ (на каждом листе):",
        "   - Линейка: для имплантов - линейка товара (AnyRidge, AnyOne и т.д.),",
        "             для протетики/лаборатории - линейка импланта (AnyRidge, AnyOne и т.д.),",
        "             для наборов/материалов - может быть пустым",
        "   - Название товара: полное название товара",
        "   - Артикул (1C): уникальный артикул товара из 1C (ОБЯЗАТЕЛЬНО для синхронизации!)",
        "   - Тип: для протетики - 'прямой' или 'угловой' (для других категорий оставить пустым)",
        "   - Диаметр: диаметр товара в мм (только для товаров с размерами, оставить пустым для наборов/материалов)",
        "   - Длина: для имплантов - длина в мм, для протетики/лаборатории - высота десны в мм",
        "   - Высота абатмента: только для протетики/лаборатории - высота абатмента в мм",
        "   - Ед. изм.: единица измерения (шт, комплект, упак и т.д.)",
        "",
        "3. ПРАВИЛА ЗАПОЛНЕНИЯ:",
        "   - Каждая строка = один товар",
        "   - Артикул (1C) должен быть уникальным и соответствовать артикулу в системе 1C",
        "   - Для имплантов: обязательно указать Линейку, Диаметр и Длину, Тип и Высота абатмента - пустые",
        "   - Для протетики/лаборатории: обязательно указать Линейку импланта, Название товара, Диаметр, Длину (высота десны) и Высоту абатмента",
        "   - Для протетики также обязательно указать Тип (прямой/угловой), для лаборатории Тип может быть пустым",
        "   - Для товаров без размеров (наборы, материалы): все размеры должны быть пустыми",
        "   - Категория и Линейка должны быть одинаковыми для товаров одной группы",
        "",
        "4. ПРИМЕРЫ:",
        "",
        "   Имплант (лист 'Импланты'):",
        "   Линейка: AnyRidge | Название: AnyRidge 3507 | Артикул: AR-3507 | Тип: (пусто) | Диаметр: 3.5 | Длина: 7.0 | Высота абатмента: (пусто) | Ед. изм.: шт",
        "",
        "   Протетика (лист 'Протетика'):",
        "   Линейка: AnyRidge | Название: EZ Post Abutment | Артикул: EP-AR-45-15-25-ST | Тип: прямой | Диаметр: 4.5 | Длина: 1.5 | Высота абатмента: 2.5 | Ед. изм.: шт",
        "",
        "   Лаборатория (лист 'Лаборатория'):",
        "   Линейка: AnyRidge | Название: Lab Analog | Артикул: LA-AR-45-15-25 | Тип: (пусто) | Диаметр: 4.5 | Длина: 1.5 | Высота абатмента: 2.5 | Ед. изм.: шт",
        "",
        "   Набор (лист 'Наборы'):",
        "   Линейка: (пусто) | Название: Anchor Kit (AO) KAGAS3001 | Артикул: KAGAS3001 | Тип: (пусто) | Диаметр: (пусто) | Длина: (пусто) | Высота абатмента: (пусто) | Ед. изм.: комплект",
        "",
        "5. ВАЖНО:",
        "   - Не удаляйте строку с заголовками!",
        "   - Не оставляйте пустые строки между товарами",
        "   - Артикул (1C) используется для синхронизации остатков с системой 1C",
        "   - После заполнения запустите: python load_catalog_from_excel.py",
        "   - Категория определяется названием листа, не нужно заполнять колонку 'Категория'",
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1)
        cell.value = instruction
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    
    ws_instructions.column_dimensions["A"].width = 100
    
    # Сохранить
    filename = "catalog_template.xlsx"
    wb.save(filename)
    print(f"Template created: {filename}")
    print("\nStructure:")
    print("  - Each sheet = one category")
    print("  - Sheets: Импланты, Протетика, Лаборатория, Наборы, материалы")
    print("\nInstructions:")
    print("1. Open catalog_template.xlsx")
    print("2. Fill in each sheet with products of that category")
    print("3. Make sure 'Артикул (1C)' column is filled for all items")
    print("4. For items with sizes: fill Diameter and Length (and Height for prosthetics)")
    print("5. For items without sizes: leave Diameter, Length and Height empty")
    print("6. After filling, run: python load_catalog_from_excel.py")

if __name__ == "__main__":
    create_catalog_template()
