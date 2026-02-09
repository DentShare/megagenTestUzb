"""
Генерация Excel шаблона для каталога продукции.

Использование:
    python generate_catalog_template.py

Создаст файл catalog_template.xlsx с колонками:
- Категория
- Линейка
- Диаметр
- Длина

Заполните таблицу и используйте load_catalog_from_excel.py для загрузки.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_catalog_template():
    """Создаёт Excel шаблон для заполнения каталога."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    
    # Заголовки
    headers = ["Категория", "Линейка", "Диаметр", "Длина", "Примечание"]
    ws.append(headers)
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры данных (можно удалить после заполнения)
    examples = [
        ["Импланты", "AnyRidge", 3.5, 8.5, "Пример строки"],
        ["Импланты", "AnyRidge", 3.5, 10.0, ""],
        ["Импланты", "AnyRidge", 3.5, 11.5, ""],
        ["Импланты", "AnyRidge", 4.0, 8.5, ""],
        ["Импланты", "AnyOne", 3.0, 10.0, ""],
        ["Наборы", "Surgical Kit", 0, 0, "Наборы без диаметра/длины"],
        ["Кость", "Bone Graft", 0, 0, "Материал без размеров"],
    ]
    
    for row in examples:
        ws.append(row)
    
    # Автоширина колонок
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    # Заморозить первую строку
    ws.freeze_panes = "A2"
    
    # Сохранить
    filename = "catalog_template.xlsx"
    wb.save(filename)
    print(f"✅ Шаблон создан: {filename}")
    print("\nИнструкция:")
    print("1. Откройте catalog_template.xlsx")
    print("2. Заполните таблицу: Категория → Линейка → Диаметр → Длина")
    print("3. Для товаров без диаметра/длины (наборы, материалы) используйте 0")
    print("4. После заполнения запустите: python load_catalog_from_excel.py")

if __name__ == "__main__":
    create_catalog_template()
