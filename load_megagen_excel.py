"""
Загрузка каталога из Megagenbot.xlsx и интеграция в бота

Структура Excel:
- Листы = категории (Импланты, Протетика, Лаборатория, Наборы, материалы)
- Первая строка блока = название линейки
- Следующие строки = товары этой линейки (до следующей линейки)
- Колонка 1 = название товара
- Колонка 2 = количество на складе
"""
import sys
import json
import re
import io
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

# Устанавливаем UTF-8 для вывода
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def extract_dimensions_from_name(name: str):
    """
    Пытается извлечь диаметр и длину из названия товара.
    
    Примеры:
    - "AnyOne 3507 C" -> diameter=3.5, length=7.0
    - "AnyOne 3510 C" -> diameter=3.5, length=10.0
    - "AnyRidge 4008" -> diameter=4.0, length=8.0
    """
    if not name:
        return None, None
    
    # Ищем паттерны типа: 3507 (3.5 диаметр, 07 длина), 4008 (4.0 диаметр, 08 длина)
    # Или: 3.5x7, 4.0x10 и т.д.
    
    # Паттерн 1: 4 цифры подряд (первые 2 = диаметр*10, последние 2 = длина*10)
    match = re.search(r'(\d{2})(\d{2})', name)
    if match:
        diam_str = match.group(1)
        length_str = match.group(2)
        try:
            diameter = float(diam_str) / 10.0
            length = float(length_str) / 10.0
            # Проверяем разумность значений (диаметр обычно 2.0-6.0, длина 5.0-20.0)
            if 2.0 <= diameter <= 6.0 and 5.0 <= length <= 20.0:
                return diameter, length
        except:
            pass
    
    # Паттерн 2: явное указание диаметра и длины (3.5x7, 4.0x10)
    match = re.search(r'(\d+\.?\d*)\s*[xх×]\s*(\d+\.?\d*)', name, re.IGNORECASE)
    if match:
        try:
            diameter = float(match.group(1))
            length = float(match.group(2))
            if 2.0 <= diameter <= 6.0 and 5.0 <= length <= 20.0:
                return diameter, length
        except:
            pass
    
    # Паттерн 3: отдельные упоминания диаметра (Ø3.5, 3.5mm) и длины (7mm, 10mm)
    diam_match = re.search(r'[øØ]?\s*(\d+\.?\d*)\s*(?:mm|мм)?', name, re.IGNORECASE)
    length_match = re.search(r'(\d+\.?\d*)\s*(?:mm|мм)', name, re.IGNORECASE)
    
    diameter = None
    length = None
    
    if diam_match:
        try:
            diameter = float(diam_match.group(1))
            if not (2.0 <= diameter <= 6.0):
                diameter = None
        except:
            pass
    
    if length_match:
        try:
            length = float(length_match.group(1))
            if not (5.0 <= length <= 20.0):
                length = None
        except:
            pass
    
    return diameter, length

def parse_sheet(ws, sheet_name: str):
    """Парсит лист Excel и возвращает структуру каталога."""
    catalog = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    current_line = None
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row or not row[0]:
            continue
        
        product_name = str(row[0]).strip()
        if not product_name:
            continue
        
        # Проверяем, является ли это названием линейки (обычно без цифр в начале или короткое название)
        # Если в строке нет количества (колонка 2 пустая или это просто название), это может быть линейка
        qty = row[1] if len(row) > 1 else None
        
        # Если колонка 2 пустая и название короткое/без артикула - это линейка
        if qty is None or qty == "":
            # Проверяем, похоже ли на название линейки (не содержит типичных паттернов артикула)
            if not re.search(r'\d{4}', product_name) and len(product_name) < 30:
                current_line = product_name
                print(f"  Found line: {current_line}")
                continue
        
        # Это товар
        if not current_line:
            current_line = "Без категории"
        
        # Пытаемся извлечь диаметр и длину
        diameter, length = extract_dimensions_from_name(product_name)
        
        if diameter is not None and length is not None:
            # Товар с размерами
            if length not in catalog[sheet_name][current_line][diameter]:
                catalog[sheet_name][current_line][diameter].append(length)
                catalog[sheet_name][current_line][diameter].sort()
        else:
            # Товар без размеров (наборы, материалы, абатменты и т.д.)
            if "no_size" not in catalog[sheet_name][current_line]:
                catalog[sheet_name][current_line]["no_size"] = []
            # Сохраняем название товара для товаров без размеров
            catalog[sheet_name][current_line]["no_size"].append(product_name)
    
    return catalog

def load_catalog_from_excel(filename: str = "Megagenbot.xlsx"):
    """Загружает каталог из Excel файла."""
    file_path = Path(filename)
    if not file_path.exists():
        print(f"ERROR: File {filename} not found!")
        return None
    
    wb = load_workbook(filename, data_only=True)
    
    print(f"Loading: {filename}")
    print(f"Sheets: {wb.sheetnames}\n")
    
    all_catalog = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name} ({ws.max_row} rows)")
        
        sheet_catalog = parse_sheet(ws, sheet_name)
        
        # Объединяем с общим каталогом
        for cat, lines in sheet_catalog.items():
            for line, diameters in lines.items():
                for key, value in diameters.items():
                    if key == "no_size":
                        if "no_size" not in all_catalog[cat][line]:
                            all_catalog[cat][line]["no_size"] = []
                        all_catalog[cat][line]["no_size"].extend(value)
                    else:
                        if key not in all_catalog[cat][line]:
                            all_catalog[cat][line][key] = []
                        all_catalog[cat][line][key].extend(value)
                        all_catalog[cat][line][key] = sorted(list(set(all_catalog[cat][line][key])))
    
    # Преобразуем в обычный dict
    result = {}
    for cat, lines in all_catalog.items():
        result[cat] = {}
        for line, diameters in lines.items():
            result[cat][line] = {}
            for key, value in diameters.items():
                if key == "no_size":
                    result[cat][line]["no_size"] = value  # Список названий товаров
                else:
                    result[cat][line][key] = sorted(list(set(value)))
    
    return result

def generate_catalog_file(catalog: dict):
    """Генерирует catalog_data.py из загруженных данных."""
    lines = [
        '"""',
        "Данные каталога продукции (загружено из Megagenbot.xlsx).",
        "Не редактировать вручную! Используйте load_megagen_excel.py",
        '"""',
        "",
        "CATALOG = {",
    ]
    
    for cat, lines_dict in sorted(catalog.items()):
        lines.append(f'    "{cat}": {{')
        for line, diameters in sorted(lines_dict.items()):
            lines.append(f'        "{line}": {{')
            # Сортируем ключи: сначала числовые (диаметры), потом строковые
            sorted_keys = sorted(
                diameters.items(),
                key=lambda x: (isinstance(x[0], str), x[0])
            )
            for key, value in sorted_keys:
                if key == "no_size":
                    if isinstance(value, list):
                        # Список товаров без размеров
                        items_str = ",\n            ".join([f'"{item}"' for item in value])
                        lines.append(f'            "no_size": [\n                {items_str}\n            ],')
                    else:
                        lines.append('            "no_size": True,')
                else:
                    lengths_str = ", ".join(str(l) for l in value)
                    lines.append(f"            {key}: [{lengths_str}],")
            lines.append("        },")
        lines.append("    },")
    
    lines.append("}")
    lines.append("")
    
    content = "\n".join(lines)
    
    with open("catalog_data.py", "w", encoding="utf-8") as f:
        f.write(content)
    
    print("\nGenerated: catalog_data.py")

def print_statistics(catalog: dict):
    """Выводит статистику по каталогу."""
    total_categories = len(catalog)
    total_lines = sum(len(lines) for lines in catalog.values())
    total_items = 0
    total_no_size = 0
    
    for cat, lines in catalog.items():
        for line, diameters in lines.items():
            for key, value in diameters.items():
                if key == "no_size":
                    if isinstance(value, list):
                        total_no_size += len(value)
                    else:
                        total_no_size += 1
                else:
                    total_items += len(value)
    
    print(f"\nStatistics:")
    print(f"  Categories: {total_categories}")
    print(f"  Lines: {total_lines}")
    print(f"  Items with sizes: {total_items}")
    print(f"  Items without sizes: {total_no_size}")
    print(f"  Total items: {total_items + total_no_size}")

def main():
    catalog = load_catalog_from_excel()
    if catalog:
        print_statistics(catalog)
        generate_catalog_file(catalog)
        
        # Также сохраняем JSON
        with open("catalog.json", "w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)
        print("Generated: catalog.json")
        
        print("\nNext step: Update keyboards/manager_kbs.py to use catalog_data.py")

if __name__ == "__main__":
    main()
