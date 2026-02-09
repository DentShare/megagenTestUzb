import csv
import json
import logging
from datetime import datetime
from io import StringIO, BytesIO
from aiogoogle import Aiogoogle
from aiogoogle.auth.creds import ServiceAccountCreds
from config import config

logger = logging.getLogger(__name__)


def export_to_xlsx(data: list[list], sheet_name: str = "Заказы") -> BytesIO:
    """
    Экспорт в Excel (.xlsx) с корректной структурой: каждый столбец и строка на своём месте.
    Удобно для фильтрации и сортировки в Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Лимит имени листа 31 символ

    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)

    # Автоширина столбцов
    if data:
        for col_idx in range(1, len(data[0]) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Закрепление первой строки (заголовок) для удобной фильтрации
    ws.freeze_panes = "A2"

    # Автофильтр по заголовкам (удобно фильтровать данные)
    if data and len(data) > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(data[0]))}{len(data)}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_csv(data: list[list]) -> BytesIO:
    """
    Экспорт в CSV. Разделитель — точка с запятой (;) для корректного открытия в Excel (русская локаль).
    Каждое значение в отдельной ячейке.
    """
    output = BytesIO()
    output.write(b'\xef\xbb\xbf')  # UTF-8 BOM для Excel
    text_buffer = StringIO()
    # Точка с запятой — стандартный разделитель для Excel в русской локали
    writer = csv.writer(text_buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL, lineterminator='\r\n')
    writer.writerows(data)
    output.write(text_buffer.getvalue().encode('utf-8'))
    output.seek(0)
    return output


async def export_to_sheets(data: list[list], sheet_name: str = "Заказы"):
    """
    Exports data to a new Google Sheet.
    Returns the URL of the sheet if successful, or XLSX buffer as fallback (корректные столбцы и строки).
    """
    # Try to use Google Sheets if credentials are configured
    if config.GOOGLE_SERVICE_ACCOUNT_JSON:
        try:
            # Parse service account credentials
            if config.GOOGLE_SERVICE_ACCOUNT_JSON.startswith('{'):
                # JSON string
                creds_dict = json.loads(config.GOOGLE_SERVICE_ACCOUNT_JSON)
            else:
                # File path
                with open(config.GOOGLE_SERVICE_ACCOUNT_JSON, 'r') as f:
                    creds_dict = json.load(f)
            
            service_account_creds = ServiceAccountCreds(**creds_dict)
            
            async with Aiogoogle(service_account_creds=service_account_creds) as aiogoogle:
                # Discover Sheets API
                sheets_v4 = await aiogoogle.discover('sheets', 'v4')
                drive_v3 = await aiogoogle.discover('drive', 'v3')
                
                # Create a new spreadsheet
                spreadsheet_title = f"Отчет по заказам {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                
                spreadsheet_body = {
                    'properties': {
                        'title': spreadsheet_title
                    },
                    'sheets': [{
                        'properties': {
                            'title': 'Заказы'
                        }
                    }]
                }
                
                # Create spreadsheet
                create_request = sheets_v4.spreadsheets.create(json=spreadsheet_body)
                create_response = await aiogoogle.as_service_account(create_request)
                spreadsheet_id = create_response['spreadsheetId']
                
                # Write data to the sheet
                range_name = 'A1'
                value_input_option = 'RAW'
                
                update_body = {
                    'values': data
                }
                
                update_request = sheets_v4.spreadsheets.values.update(
                    spreadsheetId=spreadsheet_id,
                    range=range_name,
                    valueInputOption=value_input_option,
                    json=update_body
                )
                await aiogoogle.as_service_account(update_request)
                
                # Make the sheet publicly readable (optional, or share with specific users)
                # For now, we'll just return the URL - admin can share manually
                sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
                
                logger.info(f"Successfully created Google Sheet: {sheet_url}")
                return sheet_url
                
        except Exception as e:
            logger.error(f"Failed to export to Google Sheets: {e}")
            logger.info("Falling back to XLSX export")
    
    # Fallback: XLSX — корректные столбцы и строки, удобная фильтрация в Excel
    return export_to_xlsx(data, sheet_name)

async def generate_report_data(orders):
    """
    Prepares detailed report data with manager, clinic, timestamps, and products.
    Каждая строка = один товар в заказе, что позволяет легко фильтровать в Excel.
    Headers: ID заказа, Менеджер, Клиника, Врач, Дата создания, Время создания, 
             Статус, Время сборки, Время доставки, SKU товара, Название товара, Количество
    """
    # Заголовки с раздельными столбцами для товаров
    rows = [[
        "ID заказа",
        "Менеджер",
        "Клиника",
        "Врач",
        "Дата создания",
        "Время создания",
        "Статус",
        "Время сборки",
        "Время доставки",
        "SKU товара",
        "Название товара",
        "Количество"
    ]]
    
    for o in orders:
        # Менеджер
        manager_name = o.manager.full_name if o.manager else "Неизвестен"
        
        # Клиника
        clinic_name = o.clinic.name if o.clinic else "Неизвестна"
        doctor_name = o.clinic.doctor_name if o.clinic else "-"
        
        # Дата и время создания
        created_date = o.created_at.strftime("%Y-%m-%d") if o.created_at else "-"
        created_time = o.created_at.strftime("%H:%M:%S") if o.created_at else "-"
        
        # Статус
        status = o.status.value if hasattr(o.status, 'value') else str(o.status)
        
        # Время сборки
        assembly_time = o.assembled_at.strftime("%Y-%m-%d %H:%M:%S") if o.assembled_at else "-"
        
        # Время доставки
        delivery_time = o.delivered_at.strftime("%Y-%m-%d %H:%M:%S") if o.delivered_at else "-"
        
        # Общие данные заказа (будут повторяться для каждого товара)
        order_base_data = [
            o.id,
            manager_name,
            clinic_name,
            doctor_name,
            created_date,
            created_time,
            status,
            assembly_time,
            delivery_time
        ]
        
        # Товары в заказе - каждый товар в отдельной строке
        if o.items and len(o.items) > 0:
            for item in o.items:
                rows.append([
                    *order_base_data,  # Общие данные заказа
                    item.item_sku or "-",  # SKU товара
                    item.item_name or "-",  # Название товара
                    item.quantity  # Количество
                ])
        else:
            # Если нет товаров, создаем одну строку с пустыми полями товара
            rows.append([
                *order_base_data,
                "-",  # SKU
                "Нет товаров",  # Название
                0  # Количество
            ])
    
    return rows

async def generate_product_statistics(session, start_date=None, end_date=None):
    """
    Генерирует статистику по продукции.
    Анализирует наиболее заказываемые товары за период.
    
    Returns:
        list: Список словарей с данными о товарах: {'sku', 'name', 'total_orders', 'total_quantity'}
    """
    from sqlalchemy import select, func, and_
    from database.models import Order, OrderItem
    from datetime import datetime
    
    # Базовый запрос
    stmt = (
        select(
            OrderItem.item_sku,
            OrderItem.item_name,
            func.count(func.distinct(OrderItem.order_id)).label('total_orders'),
            func.sum(OrderItem.quantity).label('total_quantity')
        )
        .join(Order, OrderItem.order_id == Order.id)
        .group_by(OrderItem.item_sku, OrderItem.item_name)
    )
    
    # Фильтрация по периоду если указан
    if start_date or end_date:
        conditions = []
        if start_date:
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, "%d.%m.%Y")
            conditions.append(Order.created_at >= start_date)
        if end_date:
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, "%d.%m.%Y")
                # Добавляем время 23:59:59 для включения всего дня
                from datetime import time
                end_date = datetime.combine(end_date.date(), time(23, 59, 59))
            conditions.append(Order.created_at <= end_date)
        
        if conditions:
            stmt = stmt.where(and_(*conditions))
    
    # Сортировка по количеству заказов и количеству единиц
    stmt = stmt.order_by(
        func.count(func.distinct(OrderItem.order_id)).desc(),
        func.sum(OrderItem.quantity).desc()
    )
    
    result = await session.execute(stmt)
    stats = []
    
    for row in result:
        stats.append({
            'sku': row.item_sku,
            'name': row.item_name,
            'total_orders': row.total_orders,
            'total_quantity': row.total_quantity
        })
    
    return stats

async def format_product_statistics(stats, limit=20):
    """
    Форматирует статистику продукции для отображения.
    
    Returns:
        list: Список строк для отображения или экспорта
    """
    if not stats:
        return [["Нет данных за выбранный период"]]
    
    # Заголовок
    rows = [["Место", "Товар", "SKU", "Кол-во заказов", "Общее кол-во (шт)"]]
    
    # Данные (топ-N)
    for idx, item in enumerate(stats[:limit], 1):
        rows.append([
            idx,
            item['name'],
            item['sku'],
            item['total_orders'],
            item['total_quantity']
        ])
    
    return rows