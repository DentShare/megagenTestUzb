"""
Загрузка каталога из Excel в таблицу catalog_items.

Использование:
    # Из скрипта (CLI):
    python -m services.catalog_loader_db [filename.xlsx]

    # Из бота (async):
    from services.catalog_loader_db import load_excel_to_db
    stats = await load_excel_to_db(session, "catalog_template.xlsx")
"""
from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Парсинг Excel
# ---------------------------------------------------------------------------

def _parse_diameter(val) -> tuple[Optional[float], Optional[float]]:
    """Извлекает (diameter, body_diameter) из строки.
    '3.5' → (3.5, None); '4.5 [3.8]' → (4.5, 3.8)."""
    if val is None:
        return (None, None)
    s = str(val).strip().replace(",", ".")
    if not s:
        return (None, None)
    part_before = s.split("[")[0].strip()
    diameter = None
    if part_before:
        m = re.search(r"([\d.]+)", part_before)
        if m:
            try:
                diameter = float(m.group(1))
            except ValueError:
                pass
    body = None
    bm = re.search(r"\[([\d.]+)]", s)
    if bm:
        try:
            body = float(bm.group(1))
        except ValueError:
            pass
    if diameter is None and body is not None:
        diameter = body
    return (diameter, body)


def _parse_show(val) -> bool:
    if val is None:
        return True
    s = str(val).strip().lower()
    return s not in ("0", "нет", "no", "дополнительно", "additional", "false", "скрыть", "hide")


def _parse_type(val, type_angles, diameter_range) -> Optional[str]:
    """Парсит product_type из ячейки Excel. Возвращает str или None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if "[n]" in s.lower():
        clean = s.replace("[N]", "").replace("[n]", "").strip()
        if clean and clean.replace(".", "").isdigit():
            return f"{int(float(clean))} [N]"
        return "[N]"
    try:
        num = float(s)
        int_val = int(num) if num == int(num) else None
        if diameter_range[0] <= num <= diameter_range[1] and (int_val is None or int_val not in type_angles):
            return None
        if int_val is not None and int_val in type_angles:
            return str(int_val)
        if num in type_angles:
            return str(int(num))
        return None
    except (ValueError, TypeError):
        sl = s.lower()
        if "прямой" in sl or "straight" in sl:
            return "0"
        if "угловой" in sl or "angled" in sl or "угол" in sl:
            return "17"
    return None


def _find_columns(headers: list[str], keywords_map: dict, skip_cols: set) -> dict:
    """Определяет индексы колонок по заголовкам."""
    col_indices: dict[str, Optional[int]] = {k: None for k in keywords_map}
    for col_type, keywords in keywords_map.items():
        if col_type in skip_cols:
            continue
        for idx, header in enumerate(headers):
            hl = header.lower()
            for kw in keywords:
                if hl == kw.lower() or kw.lower() in hl:
                    col_indices[col_type] = idx
                    break
            if col_indices[col_type] is not None:
                break
    return col_indices


def _cell(row, idx, default=None):
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return v if v is not None else default


def parse_excel(filename: str) -> List[dict]:
    """
    Парсит Excel и возвращает плоский список dict-ов
    (по одному на товар), готовых для upsert в catalog_items.
    """
    from catalog_config import (
        EXCEL_COLUMN_KEYWORDS, TYPE_ANGLES, DIAMETER_RANGE,
        SHEET_TO_CATEGORY, SKIP_SHEETS, CATEGORY_STRUCTURE,
    )

    path = Path(filename)
    if not path.exists():
        logger.error("File not found: %s", filename)
        return []

    wb = load_workbook(path, data_only=True)
    all_items: List[dict] = []

    for sheet_name in wb.sheetnames:
        sl = sheet_name.lower()
        if any(skip in sl for skip in SKIP_SHEETS):
            continue

        category = SHEET_TO_CATEGORY.get(sl)
        if not category:
            for key, cat in SHEET_TO_CATEGORY.items():
                if key in sl:
                    category = cat
                    break
        if not category:
            category = sheet_name

        ws = wb[sheet_name]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        cat_cfg = CATEGORY_STRUCTURE.get(category, {})
        skip_cols = set()
        if category in ("Наборы", "материалы"):
            skip_cols = {"product_type", "diameter", "length", "height"}

        col = _find_columns(headers, EXCEL_COLUMN_KEYWORDS, skip_cols)

        if col.get("name") is None:
            logger.warning("Sheet '%s': missing 'name' column, skipping", sheet_name)
            continue

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            name = str(_cell(row, col["name"], "")).strip()
            if not name:
                continue

            subcategory_val = str(_cell(row, col.get("category"), "")).strip() or None
            line_val = str(_cell(row, col.get("line"), "")).strip() or "Без линейки"
            sku_val = str(_cell(row, col.get("sku"), "")).strip()

            diameter, diameter_body = (None, None)
            if "diameter" not in skip_cols:
                diameter, diameter_body = _parse_diameter(_cell(row, col.get("diameter")))

            length_val = None
            if "length" not in skip_cols:
                raw = _cell(row, col.get("length"))
                if raw is not None and str(raw).strip():
                    try:
                        length_val = float(raw)
                    except (ValueError, TypeError):
                        pass

            height_val = None
            if "height" not in skip_cols:
                raw = _cell(row, col.get("height"))
                if raw is not None and str(raw).strip():
                    try:
                        height_val = float(raw)
                    except (ValueError, TypeError):
                        pass

            product_type = None
            if "product_type" not in skip_cols:
                product_type = _parse_type(
                    _cell(row, col.get("product_type")), TYPE_ANGLES, DIAMETER_RANGE
                )

            unit = str(_cell(row, col.get("unit"), "шт")).strip() or "шт"

            qty = 0
            raw_qty = _cell(row, col.get("quantity"))
            if raw_qty is not None and str(raw_qty).strip():
                try:
                    qty = max(0, int(float(raw_qty)))
                except (ValueError, TypeError):
                    pass

            show = _parse_show(_cell(row, col.get("show_immediately")))

            # Авто-генерация SKU при отсутствии
            if not sku_val:
                parts = []
                # Включаем линейку для уникальности (AnyOne vs AnyOne DEEP)
                line_short = "".join(c for c in line_val[:8] if c.isalnum()).upper()
                if line_short:
                    parts.append(line_short)
                short = "".join(c for c in name[:10] if c.isalnum() or c in "-_").upper()
                if short:
                    parts.append(short)
                if diameter is not None:
                    parts.append(f"D{int(diameter * 10)}")
                if length_val is not None:
                    parts.append(f"L{int(length_val * 10)}")
                if height_val is not None:
                    parts.append(f"H{int(height_val * 10)}")
                if product_type is not None:
                    parts.append(f"T{product_type}")
                sku_val = "-".join(parts) if parts else f"AUTO-{sheet_name}-{row_idx}"

            all_items.append({
                "category": category,
                "subcategory": subcategory_val,
                "line": line_val,
                "product_name": name,
                "product_type": product_type,
                "diameter": diameter,
                "diameter_body": diameter_body,
                "length": length_val,
                "height": height_val,
                "sku": sku_val,
                "unit": unit,
                "qty": qty,
                "show_immediately": show,
                "is_active": True,
            })

    # Дедупликация SKU: при коллизиях добавляем суффикс -2, -3, ...
    seen: dict[str, int] = {}
    for item in all_items:
        sku = item["sku"]
        if sku in seen:
            seen[sku] += 1
            item["sku"] = f"{sku}-{seen[sku]}"
        else:
            seen[sku] = 1

    logger.info("Parsed %d items from %s", len(all_items), filename)
    return all_items


# ---------------------------------------------------------------------------
# Async: загрузка в БД
# ---------------------------------------------------------------------------

async def load_excel_to_db(
    session,
    filename: str,
    deactivate_missing: bool = True,
) -> dict:
    """
    Загрузить Excel в catalog_items (upsert по SKU).
    Возвращает {"inserted": N, "updated": N, "deactivated": N, "total_parsed": N}.
    """
    from services.catalog_db import upsert_items

    items = parse_excel(filename)
    if not items:
        return {"inserted": 0, "updated": 0, "deactivated": 0, "total_parsed": 0}

    stats = await upsert_items(session, items, deactivate_missing=deactivate_missing)
    await session.commit()
    stats["total_parsed"] = len(items)

    logger.info(
        "Excel → DB: parsed=%d, inserted=%d, updated=%d, deactivated=%d",
        len(items), stats["inserted"], stats["updated"], stats["deactivated"],
    )
    return stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

async def _main_async(filename: str):
    from database.core import session_maker
    async with session_maker() as session:
        stats = await load_excel_to_db(session, filename)
        print(f"Done: {stats}")


def main():
    import asyncio
    filename = sys.argv[1] if len(sys.argv) > 1 else "catalog_template.xlsx"
    if not Path(filename).exists():
        for alt in ("Megagenbot.xlsx", "catalog.xlsx"):
            if Path(alt).exists():
                filename = alt
                break
    logging.basicConfig(level=logging.INFO)
    asyncio.run(_main_async(filename))


if __name__ == "__main__":
    main()
